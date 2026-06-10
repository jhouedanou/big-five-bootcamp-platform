# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BLACK = "0F0F0F"; GOLD = "F2B33D"; LIGHT = "F5F5F5"; WHITE = "FFFFFF"
FONT = "Arial"

thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfont(**k): return Font(name=FONT, **k)

wb = Workbook()

# ---------------------------------------------------------------- Lisez-moi
ws0 = wb.active; ws0.title = "Lisez-moi"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 110
intro = [
    ("Checklist QA — Laveiye (lot onboarding / admin / promo / webinaires / analytics)", 16, True, GOLD, BLACK),
    ("", 8, False, None, None),
    ("COMMENT UTILISER CE FICHIER", 12, True, BLACK, None),
    ("1. Importer dans Google Sheets : Drive > Nouveau > Importer > ce fichier .xlsx (ouvre en Google Sheet).", 11, False, None, None),
    ("2. Onglet « Checklist QA » : 1 ligne = 1 cas de test. Renseigner Statut, Sévérité, Bug/Notes.", 11, False, None, None),
    ("3. Colonne « Capture d'écran » : cliquer la cellule > Insertion > Image > Image dans la cellule, OU coller (Ctrl+V) une capture.", 11, False, None, None),
    ("   (Astuce Google Sheets : =IMAGE(\"URL\") si la capture est hébergée en ligne.)", 10, False, None, None),
    ("4. Statut et Sévérité ont des listes déroulantes. Onglet « Synthèse » : compteurs automatiques.", 11, False, None, None),
    ("", 8, False, None, None),
    ("LÉGENDE STATUT", 12, True, BLACK, None),
    ("À tester = pas encore vérifié   |   OK = conforme   |   KO = bug   |   Bloqué = impossible à tester   |   N/A = non applicable", 11, False, None, None),
    ("", 4, False, None, None),
    ("LÉGENDE SÉVÉRITÉ", 12, True, BLACK, None),
    ("Bloquant = empêche l'usage / paiement / accès   |   Majeur = fonction cassée mais contournable   |   Mineur = cosmétique / confort", 11, False, None, None),
    ("", 8, False, None, None),
    ("PRÉ-REQUIS AVANT DE TESTER", 12, True, BLACK, None),
    ("• Exécuter les migrations SQL #1 → #8 (voir migrations.md) dans Supabase.", 11, False, None, None),
    ("• Promo : période 01/07/2026 → 31/08/2026 — pour tester hors période, ajuster promo_campaigns.start_date/end_date.", 11, False, None, None),
    ("• Email webinaire : config Gmail API (GMAIL_*).  Analytics GA4 serveur : GA4_API_SECRET.", 11, False, None, None),
    ("• Paiement promo : produits Chariow prd_9ya1w161 (Basic) / prd_51tfnkip (Pro).", 11, False, None, None),
]
r = 1
for text, size, bold, fill, fg in intro:
    c = ws0.cell(row=r, column=1, value=text)
    c.font = hfont(size=size, bold=bold, color=(fg or BLACK))
    c.alignment = Alignment(wrap_text=True, vertical="center")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    ws0.row_dimensions[r].height = max(20, size*1.6)
    r += 1

# ---------------------------------------------------------------- Checklist
ws = wb.create_sheet("Checklist QA")
ws.sheet_view.showGridLines = False
headers = ["ID", "Module", "Fonctionnalité", "Scénario / étapes", "Résultat attendu",
           "Statut", "Sévérité", "Capture d'écran (coller ici)", "Bug / Notes",
           "Navigateur · Appareil", "Testeur", "Date"]
widths = [6, 22, 30, 52, 52, 14, 14, 46, 40, 22, 16, 14]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = hfont(size=11, bold=True, color=GOLD)
    cell.fill = PatternFill("solid", fgColor=BLACK)
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = border
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 34
ws.freeze_panes = "A2"

# (module, fonctionnalité, scénario, attendu)
cases = [
 # ----- ONBOARDING
 ("Onboarding","Redirection nouveau compte","S'inscrire puis se connecter avec un profil incomplet ; tenter d'aller sur /dashboard","Redirigé automatiquement vers /onboarding ; dashboard inaccessible"),
 ("Onboarding","Modal bloquante compte existant","Compte existant profil incomplet : naviguer dans l'app (SPA)","Modal de complétion s'affiche, contenu derrière inaccessible"),
 ("Onboarding","Modal non fermable","Sur la modal/onboarding : cliquer en dehors, croix, touche Échap","Impossible de fermer sans valider le profil"),
 ("Onboarding","Champs obligatoires","Laisser Pays / Fonction / Secteurs vides","Bouton « Valider » désactivé tant que non rempli"),
 ("Onboarding","Min 1 secteur","Valider avec 0 secteur","Refusé : au moins 1 secteur requis"),
 ("Onboarding","Max 3 secteurs","Sélectionner un 4e secteur","Sélection bloquée + toast « Vous pouvez sélectionner jusqu'à 3 secteurs maximum. »"),
 ("Onboarding","Compteur secteurs","Sélectionner 2 secteurs","Affiche « 2 / 3 secteurs sélectionnés »"),
 ("Onboarding","Recherche secteurs","Taper dans le champ recherche secteurs","Liste filtrée ; secteurs viennent de Supabase (table sectors)"),
 ("Onboarding","Fonction « Autre »","Choisir Fonction = Autre","Champ texte obligatoire apparaît ; validation exige une valeur"),
 ("Onboarding","Secteur « Autre »","Choisir le secteur Autre","Champ texte « Précisez votre secteur » apparaît"),
 ("Onboarding","Validation finale","Remplir tout et valider","profile_completed = true ; redirection vers /dashboard"),
 ("Onboarding","Tracking","Compléter l'onboarding (GA4 DebugView + table analytics_events)","onboarding_started/completed en GA4 ; onboarding_completed persisté Supabase"),
 # ----- ADMIN SEGMENTATION
 ("Admin segmentation","Accès réservé admin","Appeler /admin/audience et /api/admin/users en non-admin","403 / redirection ; aucune donnée exposée"),
 ("Admin segmentation","KPI affichés","Ouvrir /admin/audience","7 KPI : campagnes, utilisateurs actifs, marques, pays, Découverte, Basic, Pro"),
 ("Admin segmentation","Pagination serveur","Changer page et taille (10/25/50/100)","Charge page par page côté serveur ; total correct"),
 ("Admin segmentation","Filtre pays","Filtrer par un pays","Liste filtrée ; compteur « X utilisateurs trouvés » à jour"),
 ("Admin segmentation","Filtres combinables","Pays + plan + statut d'activité + tag ensemble","Résultats respectent TOUS les filtres simultanément"),
 ("Admin segmentation","Statut d'activité","Vérifier un user jamais connecté / actif / dormant","Libellé correct (Jamais connecté / Actif récent / Actif / Inactif / Dormant)"),
 ("Admin segmentation","Type d'accès (codes)","Vérifier colonne Type d'accès","Affiche libellés (Accès payant, Bêta testeur, Non abonné…) — pas de code brut"),
 ("Admin segmentation","Plan sans « free »","User sans abonnement","Plan affiché « Non abonné » ; jamais « free »"),
 ("Admin segmentation","Colonnes obligatoires","Vérifier l'entête du tableau","Téléphone, Dernière connexion, Dernière activité, Statut activité/abonnement présents"),
 ("Admin segmentation","Réinitialiser filtres","Cliquer « Réinitialiser les filtres »","Tous les filtres remis à zéro ; liste complète"),
 ("Admin segmentation","Empty state","Filtrer pour 0 résultat","Affiche « Aucun utilisateur ne correspond à ces filtres. »"),
 ("Admin segmentation","Créer un tag","Créer un tag (nom + couleur)","Tag créé, visible dans la liste des tags"),
 ("Admin segmentation","Tag 1 utilisateur","Appliquer un tag via le menu d'une ligne","Badge tag apparaît sur la ligne"),
 ("Admin segmentation","Tag multi (checkbox)","Cocher plusieurs users + appliquer un tag","Tag appliqué à tous les sélectionnés"),
 ("Admin segmentation","Tag tous résultats filtrés","Sélectionner tous les résultats du filtre + appliquer","Confirmation « …appliquer le tag [X] à [N] utilisateurs » puis appliqué à tout le filtre"),
 # ----- PROMO
 ("Promo","Bannière visible","User non-Pro connecté pendant la période promo","Bannière promo affichée en haut avec compte à rebours qui défile"),
 ("Promo","Bannière masquée (Pro actif)","Se connecter en Pro actif","Bannière promo NON affichée"),
 ("Promo","Bannière masquée (expirée)","Mettre end_date dans le passé","Bannière disparaît automatiquement"),
 ("Promo","CTA bannière","Cliquer « Voir les offres »","Redirige vers /checkout?source=promo_banner"),
 ("Promo","Popup 1×/jour","Se connecter (1re fois du jour)","Popup central s'affiche une seule fois ; pas à chaque navigation"),
 ("Promo","Popup fermeture","Fermer via croix et via « Plus tard »","Popup se ferme ; ne réapparaît pas le même jour"),
 ("Promo","CTA popup","Cliquer « Voir les offres » dans le popup","Redirige vers /checkout?source=promo_popup"),
 ("Promo","Checkout offres","Ouvrir /checkout","Offre normale Pro + 2 promos ; Découverte absent"),
 ("Promo","Badge + compte à rebours","Voir les cartes promo","Badge « Offre limitée » + compte à rebours visibles"),
 ("Promo","Récap dynamique","Sélectionner une offre","Récap (choix, plan, durée, prix, total) mis à jour"),
 ("Promo","Paiement Chariow","Choisir Promo Pro, renseigner pays+téléphone, payer","Redirection Chariow (produit prd_51tfnkip), montant 10 000 FCFA"),
 ("Promo","Activation par webhook","Finaliser un paiement de test","Abonnement activé via webhook (PAS par retour front) ; plan/fin à jour"),
 # ----- WEBINAIRES
 ("Webinaires","Admin créer session","Admin /admin/webinars > Nouvelle session","Session créée (statut Brouillon par défaut)"),
 ("Webinaires","Publier / dépublier","Passer le statut Publié puis Brouillon","Statut change ; apparaît/disparaît de /webinaires"),
 ("Webinaires","Terminer / annuler","Passer Terminé puis Annulé","Statut public reflété (Terminé / Annulé)"),
 ("Webinaires","Activer/désactiver inscriptions","Toggler « Inscriptions »","Inscription possible/impossible côté /webinaires"),
 ("Webinaires","Lien public partageable","Copier le lien public d'une session","URL /webinaires/[slug]/preview copiée"),
 ("Webinaires","Page /webinaires alimentée","Publier une session puis ouvrir /webinaires","Session apparaît automatiquement (non statique) ; navbar présente"),
 ("Webinaires","Prochaine en haut + badges","Voir le programme","Prochaine session en premier ; badge de statut sur chaque carte"),
 ("Webinaires","Bloc dashboard","Ouvrir le dashboard","Bloc « À ne pas manquer » avec prochaine session + lien programme"),
 ("Webinaires","Inscription","Cliquer « S'inscrire » (connecté)","Inscription enregistrée + toast succès + proposition calendrier"),
 ("Webinaires","Anti-doublon","Se réinscrire à la même session","Refusé : déjà inscrit (pas de doublon)"),
 ("Webinaires","Email confirmation","S'inscrire et vérifier la boîte mail","Email « Confirmation d'inscription — Session #BigFiveDécrypte » reçu"),
 ("Webinaires","Ajout calendrier","Tester Google / Outlook / .ics","Les 3 options génèrent le bon événement (titre, date, heure, lien)"),
 ("Webinaires","Aperçu public","Ouvrir /webinaires/[slug]/preview déconnecté","Aperçu (titre, date, heure, desc courte) sans interface connectée"),
 ("Webinaires","Redirection intention","Depuis l'aperçu, cliquer « Se connecter » puis se connecter","Revient vers le programme ; la session est mise en évidence (halo)"),
 # ----- ANALYTICS
 ("Analytics","Events UI → GA4","Déclencher promo_banner_clicked, checkout_option_selected… (GA4 DebugView)","Events visibles dans GA4 avec page_url"),
 ("Analytics","Persistance Supabase","Déclencher un event critique","Ligne dans analytics_events (user_id, source, page_url, metadata)"),
 ("Analytics","Measurement Protocol","Avec GA4_API_SECRET, faire un paiement/inscription","Event critique remonte aussi côté serveur GA4 (MP)"),
 ("Analytics","Robustesse","Couper le réseau analytics / mauvaise config","L'app continue de fonctionner (tracking best-effort, jamais bloquant)"),
 # ----- MODELE ACTIVITE / ACCES
 ("Modèle accès","last_activity_at maj","Se connecter puis faire une action trackée","last_activity_at mis à jour (login + activity events)"),
 ("Modèle accès","KPI utilisateurs actifs","Comparer KPI avec activité réelle 30j","Compte = activité (events 30j) ∪ dernière activité < 30j"),
 ("Modèle accès","Aucun plan « free »","Parcourir UI admin + tableau","Le terme « free » n'apparaît nulle part ; null = Non abonné"),
 # ----- TRANSVERSAL
 ("Transversal","Responsive mobile","Tester onboarding / webinaires / checkout / admin sur mobile","Mise en page correcte ; filtres repliables en panneau sur mobile"),
 ("Transversal","États de chargement","Charger pages avec données lentes","Skeleton/loader affiché ; pas d'écran vide cassé"),
 ("Transversal","États erreur","Forcer une erreur API","Message d'erreur propre + possibilité de réessayer"),
]

data_font = hfont(size=10)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(wrap_text=True, vertical="center", horizontal="center")
mod_colors = {}
palette = ["FDECC8","E7F0FD","E9F7EF","FBE9E7","F3E8FD","FFF4D6","EAEDED"]
for idx, (mod, *_rest) in enumerate(cases):
    if mod not in mod_colors:
        mod_colors[mod] = palette[len(mod_colors) % len(palette)]

row = 2
for i, (mod, feat, scen, exp) in enumerate(cases, start=1):
    vals = [f"T{i:02d}", mod, feat, scen, exp, "À tester", "", "", "", "", "", ""]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = data_font
        c.border = border
        c.alignment = center if col in (1,6,7) else wrap
        if col == 2:
            c.fill = PatternFill("solid", fgColor=mod_colors[mod])
    ws.row_dimensions[row].height = 80
    row += 1

last = row - 1

# Listes déroulantes
dv_stat = DataValidation(type="list", formula1='"À tester,OK,KO,Bloqué,N/A"', allow_blank=True)
dv_sev = DataValidation(type="list", formula1='"Bloquant,Majeur,Mineur,—"', allow_blank=True)
ws.add_data_validation(dv_stat); ws.add_data_validation(dv_sev)
dv_stat.add(f"F2:F{last}"); dv_sev.add(f"G2:G{last}")

# ---------------------------------------------------------------- Synthèse
ws2 = wb.create_sheet("Synthèse")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 14
t = ws2.cell(row=1, column=1, value="Synthèse QA")
t.font = hfont(size=14, bold=True, color=BLACK)
rng = f"'Checklist QA'!F2:F{last}"
rows = [
    ("Total cas de test", f'=COUNTA(\'Checklist QA\'!A2:A{last})'),
    ("À tester", f'=COUNTIF({rng},"À tester")'),
    ("OK", f'=COUNTIF({rng},"OK")'),
    ("KO", f'=COUNTIF({rng},"KO")'),
    ("Bloqué", f'=COUNTIF({rng},"Bloqué")'),
    ("N/A", f'=COUNTIF({rng},"N/A")'),
    ("Bugs bloquants", f'=COUNTIF(\'Checklist QA\'!G2:G{last},"Bloquant")'),
    ("Bugs majeurs", f'=COUNTIF(\'Checklist QA\'!G2:G{last},"Majeur")'),
    ("Bugs mineurs", f'=COUNTIF(\'Checklist QA\'!G2:G{last},"Mineur")'),
    ("% OK", f'=IF(B2=0,0,B4/B2)'),
]
rr = 2
for label, formula in rows:
    a = ws2.cell(row=rr, column=1, value=label); a.font = hfont(size=11)
    b = ws2.cell(row=rr, column=2, value=formula); b.font = hfont(size=11, bold=True)
    if label == "% OK":
        b.number_format = "0.0%"
    rr += 1

import os
out = os.path.join(os.path.dirname(__file__), "checklist-qa-laveiye.xlsx")
wb.save(out)
print("SAVED", out)
